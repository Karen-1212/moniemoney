#!/usr/bin/env python3
"""Consumer Discretionary PE Portfolio Factsheet (Excel) — test period 2025+."""

from __future__ import annotations

import io
import sys
import warnings
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src import DEFAULT_SECTOR, TEST_START as TEST_START_STR, TRAIN_DATA_START, TRAIN_END as TRAIN_END_STR, sector_dir

warnings.filterwarnings("ignore", category=FutureWarning)

# ---------------------------------------------------------------------------
# Colour scheme (pink / purple / blue) — dark text for legibility
# ---------------------------------------------------------------------------
C_HEADER_BG = "6A1B9A"
C_HEADER_FG = "FFFFFF"
C_ACCENT = "7E57C2"
# Distinct chart lines: vivid magenta vs deep blue
C_STRAT_LINE = "#E91E8C"
C_BENCH_LINE = "#0D47A1"
C_ROW_ALT = "F3E5F5"
C_ROW_WHITE = "FFFFFF"
C_SECTION = "EDE7F6"
C_TEXT = "1A237E"
C_MUTED = "4527A0"
C_GRID = "#E1BEE7"
C_ZERO = "#B39DDB"

FINANCE_RATE = 0.04  # flat 4% for leveraged books (label >= 1x)
TX_COST = 0.01  # flat 1%
# Column labels 0x..3x map to exposure multipliers 1..4 (0x = unlevered strategy)
LEVERAGE_LEVELS = (0, 1, 2, 3)
TEST_START = pd.Timestamp(TEST_START_STR)
TRAIN_START = TRAIN_DATA_START
TRAIN_END = TRAIN_END_STR

PORTFOLIO_BLURB = (
    "About this portfolio: A dollar-neutral relative-PE pairs book in S&P 500 Consumer Discretionary. "
    "Pairs are restricted to the same GICS sub-industry, screened for significant PE co-movement, "
    "then ranked by mean-reversion score (correlation, ADF, half-life, excursion/revert stats). "
    "Up to 10 pairs may trade; sector capital is split equally among open trades and rebalanced on "
    "entry/exit. Built for sophisticated allocators seeking low-beta, market-neutral style exposure "
    "within consumer discretionary—complementary to long-only sector beta, not a replacement for it."
)


def load_equity(xlsx: Path) -> pd.DataFrame:
    eq = pd.read_excel(xlsx, sheet_name="Equity")
    df = eq.iloc[:, :3].copy()
    df.columns = ["date", "strategy_value", "benchmark_value"]
    df["date"] = pd.to_datetime(df["date"])
    df = df.dropna().sort_values("date").reset_index(drop=True)
    return df


def load_trades(xlsx: Path) -> pd.DataFrame:
    try:
        tr = pd.read_excel(xlsx, sheet_name="Trades")
    except Exception:
        return pd.DataFrame()
    if tr.empty:
        return tr
    for col in ("open_date", "close_date"):
        if col in tr.columns:
            tr[col] = pd.to_datetime(tr[col], errors="coerce")
    return tr


def slice_test_period(df: pd.DataFrame, start: pd.Timestamp = TEST_START) -> pd.DataFrame:
    """Keep test window and rebase both series to $100k at the first test date."""
    out = df.loc[df["date"] >= start].copy().reset_index(drop=True)
    if out.empty:
        return out
    s0 = float(out["strategy_value"].iloc[0])
    b0 = float(out["benchmark_value"].iloc[0])
    capital = 100_000.0
    out["strategy_value"] = capital * out["strategy_value"].astype(float) / s0
    out["benchmark_value"] = capital * out["benchmark_value"].astype(float) / b0
    return out


def daily_returns(values: pd.Series) -> pd.Series:
    return values.astype(float).pct_change().dropna()


def ann_return(r: pd.Series) -> float:
    if r.empty:
        return 0.0
    total = float((1.0 + r).prod() - 1.0)
    years = len(r) / 252.0
    if years <= 0:
        return 0.0
    return float((1.0 + total) ** (1.0 / years) - 1.0)


def max_drawdown(equity: pd.Series) -> float:
    peak = equity.cummax()
    return float((equity / peak - 1.0).min())


def tracking_error(strat: pd.Series, bench: pd.Series) -> float:
    common = strat.index.intersection(bench.index)
    excess = strat.loc[common] - bench.loc[common]
    return float(excess.std(ddof=1) * np.sqrt(252.0)) if len(excess) > 2 else np.nan


def beta_r2(strat: pd.Series, bench: pd.Series) -> tuple[float, float]:
    common = strat.index.intersection(bench.index)
    y = strat.loc[common].to_numpy(dtype=float)
    x = bench.loc[common].to_numpy(dtype=float)
    if len(x) < 3 or np.std(x) == 0:
        return np.nan, np.nan
    x_ = np.column_stack([np.ones(len(x)), x])
    coef, _, _, _ = np.linalg.lstsq(x_, y, rcond=None)
    yhat = x_.dot(coef)
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - y.mean()) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
    return float(coef[1]), float(r2)


def compute_metrics(strat_r: pd.Series, bench_r: pd.Series, strat_eq: pd.Series, bench_eq: pd.Series) -> dict:
    import pyfolio as pf

    sstats = pf.timeseries.perf_stats(strat_r, factor_returns=bench_r)
    bstats = pf.timeseries.perf_stats(bench_r)
    beta, r2 = beta_r2(strat_r, bench_r)
    te = tracking_error(strat_r, bench_r)

    def g(stats, key):
        return float(stats[key]) if key in stats.index else np.nan

    return {
        "Sharpe ratio": (g(sstats, "Sharpe ratio"), g(bstats, "Sharpe ratio")),
        "Sortino ratio": (g(sstats, "Sortino ratio"), g(bstats, "Sortino ratio")),
        "Volatility (ann.)": (g(sstats, "Annual volatility"), g(bstats, "Annual volatility")),
        "Tracking error": (te, np.nan),
        "Daily VaR (95%)": (g(sstats, "Daily value at risk"), g(bstats, "Daily value at risk")),
        "Beta": (beta, 1.0),
        "Max drawdown": (max_drawdown(strat_eq), max_drawdown(bench_eq)),
        "R squared": (r2, np.nan),
    }


def leverage_table(strat_ann: float, bench_ann: float) -> pd.DataFrame:
    """
    Column Lx uses exposure multiplier (L + 1).
    So 0x = unlevered strategy (1× returns), 1x = 2×, …, 4x = 5×.
    """
    labels = [
        "Annualised return (%)",
        "Finance cost (%)",
        "Transaction costs (%)",
        "Net returns (%)",
        "Benchmark return (%)",
        "Alpha (net − benchmark) (%)",
    ]
    data = {lab: [] for lab in labels}
    for L in LEVERAGE_LEVELS:
        mult = L + 1  # 0x -> 1× exposure
        ann = mult * strat_ann
        # 0x = unlevered (no finance cost); L >= 1 is leveraged → flat 4%
        fin = FINANCE_RATE if L >= 1 else 0.0
        tx = TX_COST
        bench_L = mult * bench_ann
        net = ann - fin - tx
        alpha = net - bench_L
        data["Annualised return (%)"].append(ann * 100.0)
        data["Finance cost (%)"].append(fin * 100.0)
        data["Transaction costs (%)"].append(tx * 100.0)
        data["Net returns (%)"].append(net * 100.0)
        data["Benchmark return (%)"].append(bench_L * 100.0)
        data["Alpha (net − benchmark) (%)"].append(alpha * 100.0)
    return pd.DataFrame(
        {f"{L}x": [data[lab][i] for lab in labels] for i, L in enumerate(LEVERAGE_LEVELS)},
        index=labels,
    )


def render_value_chart(df: pd.DataFrame) -> bytes:
    dates = pd.to_datetime(df["date"])
    strat_v = df["strategy_value"].astype(float)
    bench_v = df["benchmark_value"].astype(float)
    # Cumulative return (%) vs test-start rebase
    strat = (strat_v / float(strat_v.iloc[0]) - 1.0) * 100.0
    bench = (bench_v / float(bench_v.iloc[0]) - 1.0) * 100.0

    fig, ax = plt.subplots(figsize=(9.2, 4.6), dpi=150)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#FDFBFF")

    ax.plot(dates, strat, color=C_STRAT_LINE, linewidth=2.4, label="Relative-PE Strategy")
    ax.plot(
        dates,
        bench,
        color=C_BENCH_LINE,
        linewidth=2.2,
        linestyle="-",
        label="S&P 500 Consumer Discretionary",
    )

    ax.set_title(
        "PERFORMANCE HISTORY (Test Period)",
        fontsize=13,
        fontweight="bold",
        color="#6A1B9A",
        loc="left",
        pad=10,
    )
    ax.set_xlabel("Date", fontsize=10, color="#311B92")
    ax.set_ylabel("Cumulative return (%)", fontsize=10, color="#311B92")
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}%"))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    fig.autofmt_xdate(rotation=0, ha="center")

    # Place 0% at 40% of the axis height from the bottom:
    # (0 - y_lo) / (y_hi - y_lo) = 0.40  =>  y_lo = -(2/3) * y_hi
    data_min = float(min(strat.min(), bench.min()))
    data_max = float(max(strat.max(), bench.max()))
    y_hi = max(data_max, 0.0)
    if data_min < 0:
        # Need y_lo <= data_min with y_lo = -(2/3)*y_hi  =>  y_hi >= -1.5 * data_min
        y_hi = max(y_hi, -1.5 * data_min)
    if y_hi <= 0:
        y_hi = max(abs(data_min) * 1.5, 1.0)
    y_lo = -(0.4 / 0.6) * y_hi
    # Keep a tiny headroom above the highest point if it sits on the rim
    if data_max > 0 and data_max >= y_hi * 0.98:
        y_hi = data_max / 0.98
        y_lo = -(0.4 / 0.6) * y_hi
        if data_min < y_lo:
            y_hi = max(y_hi, -1.5 * data_min)
            y_lo = -(0.4 / 0.6) * y_hi
    ax.set_ylim(y_lo, y_hi)

    ax.axhline(0.0, color=C_ZERO, linewidth=1.0, linestyle="-", alpha=0.9)
    ax.grid(True, which="major", color=C_GRID, linewidth=0.8)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color("#7E57C2")

    ax.tick_params(colors="#4527A0", labelsize=9)
    leg = ax.legend(loc="upper left", frameon=True, fontsize=9)
    leg.get_frame().set_edgecolor("#CE93D8")
    leg.get_frame().set_facecolor("white")

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _border() -> Border:
    thin = Side(style="thin", color="B39DDB")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_trades_sheet(wb: Workbook, trades: pd.DataFrame) -> None:
    ws = wb.create_sheet("Trades")
    ws["A1"] = "Consumer Discretionary PE — Test-Period Trades"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=C_HEADER_FG)
    ws["A1"].fill = _fill(C_HEADER_BG)
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 24

    if trades.empty:
        ws["A3"] = "No trades in the test period."
        return

    display = trades.copy()
    # Prefer a stable column order when present
    preferred = [
        "pair",
        "symbol_a",
        "symbol_b",
        "score",
        "open_date",
        "close_date",
        "direction",
        "entry_z",
        "exit_z",
        "hold_days",
        "slot_notional",
        "pnl_usd",
        "return",
    ]
    cols = [c for c in preferred if c in display.columns] + [c for c in display.columns if c not in preferred]
    display = display[cols]

    for r_idx, row in enumerate(dataframe_to_rows(display, index=False, header=True), 3):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _border()
            if r_idx == 3:
                cell.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10)
                cell.fill = _fill(C_HEADER_BG)
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            else:
                cell.font = Font(name="Calibri", size=9, color=C_TEXT)
                cell.fill = _fill(C_ROW_ALT if (r_idx % 2) else C_ROW_WHITE)
                col_name = cols[c_idx - 1]
                if isinstance(val, (int, float)) and np.isfinite(val):
                    if col_name in ("return",):
                        cell.number_format = "0.00%"
                    elif col_name in ("pnl_usd", "slot_notional"):
                        cell.number_format = "#,##0.00"
                    elif col_name in ("entry_z", "exit_z", "score"):
                        cell.number_format = "0.000"
                    elif col_name == "hold_days":
                        cell.number_format = "0"

    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(22, len(str(col)) + 4))

    last_col = get_column_letter(len(cols))
    ws.auto_filter.ref = f"A3:{last_col}{2 + len(display)}"
    ws.freeze_panes = "A4"


def write_factsheet(
    path: Path,
    df: pd.DataFrame,
    lev: pd.DataFrame,
    metrics: dict,
    chart_png: bytes,
    trades: pd.DataFrame,
    period_start: str,
    period_end: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Factsheet"

    widths = {"A": 3, "B": 34, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B2:H2")
    title = ws["B2"]
    title.value = "Consumer Discretionary Portfolio Factsheet"
    title.font = Font(name="Calibri", size=20, bold=True, color=C_HEADER_FG)
    title.fill = _fill(C_HEADER_BG)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 32

    ws.merge_cells("B3:H3")
    sub = ws["B3"]
    sub.value = (
        f"Relative-PE Strategy vs S&P 500 Consumer Discretionary  |  "
        f"Train: {TRAIN_START} → {TRAIN_END}  |  "
        f"Test: {period_start} → {period_end}  |  Rebased to $100,000 at test start"
    )
    sub.font = Font(name="Calibri", size=10, color=C_MUTED)
    sub.fill = _fill(C_SECTION)
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    ws.merge_cells("B4:H5")
    about = ws["B4"]
    about.value = PORTFOLIO_BLURB
    about.font = Font(name="Calibri", size=9, color=C_TEXT)
    about.fill = _fill("F8F5FC")
    about.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 28

    img_path = path.parent / "_factsheet_chart_tmp.png"
    img_path.write_bytes(chart_png)
    img = XLImage(str(img_path))
    img.width = 620
    img.height = 310
    ws.add_image(img, "B7")

    for r in range(7, 24):
        ws.row_dimensions[r].height = 14

    lev_start = 25
    ws.merge_cells(f"B{lev_start}:F{lev_start}")
    h = ws[f"B{lev_start}"]
    h.value = "Leverage Scenario Analysis (test-period annualised returns)"
    h.font = Font(name="Calibri", size=13, bold=True, color=C_HEADER_FG)
    h.fill = _fill(C_ACCENT)
    h.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[lev_start].height = 22

    note_row = lev_start + 1
    ws.merge_cells(f"B{note_row}:F{note_row}")
    ncell = ws[f"B{note_row}"]
    ncell.value = (
        "0x = unlevered strategy (1× gross return). 1x/2x/3x = 2×/3×/4× gross return.  "
        "Finance cost: flat 4% p.a. when leverage ≥ 1x (0% at 0x).  "
        "Transaction costs: flat 1% p.a. on all columns.  "
        f"Pair selection / training window: {TRAIN_START} → {TRAIN_END}."
    )
    ncell.font = Font(name="Calibri", size=8, italic=True, color=C_MUTED)
    ncell.fill = _fill(C_SECTION)

    header_row = lev_start + 2
    ws.cell(row=header_row, column=2, value="Metric").font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10)
    ws.cell(row=header_row, column=2).fill = _fill(C_HEADER_BG)
    ws.cell(row=header_row, column=2).border = _border()
    for j, col in enumerate(lev.columns, start=3):
        cell = ws.cell(row=header_row, column=j, value=col)
        cell.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center")
        cell.border = _border()

    for i, (idx, row) in enumerate(lev.iterrows()):
        r = header_row + 1 + i
        fill = _fill(C_ROW_ALT if i % 2 else C_ROW_WHITE)
        c0 = ws.cell(row=r, column=2, value=idx)
        c0.font = Font(name="Calibri", size=10, color=C_TEXT, bold=(idx.startswith("Net") or idx.startswith("Alpha")))
        c0.fill = fill
        c0.border = _border()
        for j, col in enumerate(lev.columns, start=3):
            cell = ws.cell(row=r, column=j, value=float(row[col]))
            cell.number_format = "0.00"
            cell.font = Font(
                name="Calibri",
                size=10,
                color=C_TEXT,
                bold=(idx.startswith("Net") or idx.startswith("Alpha")),
            )
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()

    met_start = header_row + len(lev) + 3
    ws.merge_cells(f"B{met_start}:E{met_start}")
    mh = ws[f"B{met_start}"]
    mh.value = "Key Portfolio Metrics (test period)"
    mh.font = Font(name="Calibri", size=13, bold=True, color=C_HEADER_FG)
    mh.fill = _fill(C_ACCENT)
    mh.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[met_start].height = 22

    mhdr = met_start + 1
    for col, label in enumerate(["Metric", "Strategy", "Benchmark", "Difference"], start=2):
        cell = ws.cell(row=mhdr, column=col, value=label)
        cell.font = Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center" if col > 2 else "left")
        cell.border = _border()

    pct_metrics = {"Volatility (ann.)", "Tracking error", "Daily VaR (95%)", "Max drawdown"}
    for i, (name, (s_val, b_val)) in enumerate(metrics.items()):
        r = mhdr + 1 + i
        fill = _fill(C_ROW_ALT if i % 2 else C_ROW_WHITE)
        ws.cell(row=r, column=2, value=name).font = Font(name="Calibri", size=10, color=C_TEXT)
        ws.cell(row=r, column=2).fill = fill
        ws.cell(row=r, column=2).border = _border()

        for col, val in ((3, s_val), (4, b_val)):
            cell = ws.cell(
                row=r,
                column=col,
                value=None if val is None or (isinstance(val, float) and not np.isfinite(val)) else float(val),
            )
            cell.number_format = "0.00%" if name in pct_metrics else "0.00"
            cell.font = Font(name="Calibri", size=10, color=C_TEXT)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = _border()

        diff = np.nan
        if np.isfinite(s_val) and np.isfinite(b_val):
            diff = float(s_val) - float(b_val)
        dcell = ws.cell(row=r, column=5, value=None if not np.isfinite(diff) else float(diff))
        dcell.number_format = "0.00%" if name in pct_metrics else "0.00"
        dcell.font = Font(name="Calibri", size=10, color=C_TEXT)
        dcell.fill = fill
        dcell.alignment = Alignment(horizontal="center")
        dcell.border = _border()

    foot = mhdr + len(metrics) + 2
    ws.merge_cells(f"B{foot}:F{foot}")
    fcell = ws[f"B{foot}"]
    fcell.value = (
        f"Notes: Training / pair-selection window {TRAIN_START} → {TRAIN_END}; "
        f"performance above is out-of-sample test {period_start} → {period_end}. "
        "Dynamic equal capital among active trades (max 10). "
        "Beta & R² from OLS of daily strategy returns on benchmark. "
        "VaR is daily 95% (pyfolio). Full trade list on the Trades sheet."
    )
    fcell.font = Font(name="Calibri", size=8, italic=True, color=C_MUTED)
    fcell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[foot].height = 40

    write_trades_sheet(wb, trades)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    try:
        img_path.unlink(missing_ok=True)
    except Exception:
        pass


def main() -> None:
    xlsx = sector_dir(DEFAULT_SECTOR) / "portfolio_backtest.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"Missing {xlsx}. Run scripts/portfolio_backtest.py first.")

    df_raw = load_equity(xlsx)
    trades = load_trades(xlsx)
    df = slice_test_period(df_raw, TEST_START)
    if df.empty:
        raise SystemExit(f"No equity rows on/after {TEST_START.date()} in {xlsx}")

    if not trades.empty and "open_date" in trades.columns:
        trades = trades.loc[trades["open_date"] >= TEST_START].reset_index(drop=True)

    strat_eq = df.set_index("date")["strategy_value"].astype(float)
    bench_eq = df.set_index("date")["benchmark_value"].astype(float)
    strat_r = daily_returns(strat_eq)
    bench_r = daily_returns(bench_eq)
    common = strat_r.index.intersection(bench_r.index)
    strat_r = strat_r.loc[common]
    bench_r = bench_r.loc[common]

    strat_ann = ann_return(strat_r)
    bench_ann = ann_return(bench_r)
    lev = leverage_table(strat_ann, bench_ann)
    metrics = compute_metrics(strat_r, bench_r, strat_eq.loc[common.min() :], bench_eq.loc[common.min() :])
    chart = render_value_chart(df)

    out = sector_dir(DEFAULT_SECTOR) / "Consumer_Discretionary_Portfolio_Factsheet.xlsx"
    write_factsheet(
        out,
        df,
        lev,
        metrics,
        chart,
        trades,
        period_start=df["date"].min().strftime("%Y-%m-%d"),
        period_end=df["date"].max().strftime("%Y-%m-%d"),
    )
    print(f"Source: {xlsx} (sliced from {TEST_START.date()})")
    print(f"Wrote {out}")
    print(f"Test period: {df['date'].min().date()} → {df['date'].max().date()}")
    print(f"Strategy ann. return: {strat_ann:.2%} | Benchmark ann. return: {bench_ann:.2%}")
    print(f"Trades listed: {len(trades)}")
    print(lev.round(2).to_string())


if __name__ == "__main__":
    main()
