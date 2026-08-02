#!/usr/bin/env python3
"""Pyfolio (pyfolio-reloaded) analysis of the relative-PE portfolio backtest."""

from __future__ import annotations

import argparse
import os
import sys
import warnings
from pathlib import Path

# Writable matplotlib cache (sandbox / restricted home dirs)
ROOT = Path(__file__).resolve().parents[1]
os.environ.setdefault("MPLCONFIGDIR", str(ROOT / ".mplconfig"))
(ROOT / ".mplconfig").mkdir(exist_ok=True)

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, str(ROOT))
from src import DEFAULT_SECTOR, sector_dir

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=r'.*zipline.*')

BENCH_LABEL = "S&P 500 Consumer Discretionary"


def load_returns_from_excel(
    xlsx: Path, bench_label: str = BENCH_LABEL
) -> tuple[pd.Series, pd.Series]:
    """Daily simple returns for strategy and sector benchmark."""
    wb = load_workbook(xlsx, data_only=True)
    if "Equity" not in wb.sheetnames:
        raise SystemExit(f"No Equity sheet in {xlsx}")
    ws = wb["Equity"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # Columns: date, strategy_value, benchmark_value, ...
    df = pd.DataFrame(
        rows,
        columns=[
            "date",
            "strategy_value",
            "benchmark_value",
            "strategy_cumret",
            "benchmark_cumret",
        ],
    )
    df["date"] = pd.to_datetime(df["date"])
    df = df.dropna(subset=["strategy_value", "benchmark_value"]).set_index("date").sort_index()
    if getattr(df.index, "tz", None) is not None:
        df.index = df.index.tz_localize(None)

    returns = df["strategy_value"].pct_change().dropna()
    benchmark = df["benchmark_value"].pct_change().dropna()
    common = returns.index.intersection(benchmark.index)
    returns = returns.loc[common].astype(float)
    benchmark = benchmark.loc[common].astype(float)
    returns.name = "strategy"
    benchmark.name = bench_label
    return returns, benchmark


def compute_stats(returns: pd.Series, benchmark: pd.Series) -> pd.DataFrame:
    import pyfolio as pf

    bench_label = benchmark.name or "benchmark"
    stats = pf.timeseries.perf_stats(returns, factor_returns=benchmark)
    stats = stats.rename("strategy").to_frame()
    bench_stats = pf.timeseries.perf_stats(benchmark)
    stats[bench_label] = bench_stats.reindex(stats.index)
    return stats


def save_tear_sheet_figures(
    returns: pd.Series,
    benchmark: pd.Series,
    out_dir: Path,
) -> list[Path]:
    """Generate pyfolio returns tear sheet and key plots; return saved image paths."""
    import pyfolio as pf

    out_dir.mkdir(parents=True, exist_ok=True)
    saved: list[Path] = []

    # Full returns tear sheet figure (when supported)
    try:
        fig = pf.create_returns_tear_sheet(
            returns,
            benchmark_rets=benchmark,
            return_fig=True,
        )
        if fig is not None:
            path = out_dir / "pyfolio_returns_tearsheet.png"
            fig.savefig(path, dpi=140, bbox_inches="tight", facecolor="white")
            plt.close(fig)
            saved.append(path)
            print(f"Wrote {path}")
    except Exception as exc:
        print(f"create_returns_tear_sheet failed ({exc}); writing individual plots instead")

    # Always write a few standalone plots for reliability
    plotters = [
        ("cumulative_returns.png", lambda: pf.plotting.plot_rolling_returns(returns, factor_returns=benchmark)),
        ("drawdown.png", lambda: pf.plotting.plot_drawdown_periods(returns, top=5)),
        ("monthly_heatmap.png", lambda: pf.plotting.plot_monthly_returns_heatmap(returns)),
        ("rolling_sharpe.png", lambda: pf.plotting.plot_rolling_sharpe(returns)),
        ("returns_dist.png", lambda: pf.plotting.plot_returns(returns)),
    ]

    for filename, plot_fn in plotters:
        try:
            plt.close("all")
            ax_or_fig = plot_fn()
            fig = plt.gcf()
            if ax_or_fig is not None and hasattr(ax_or_fig, "figure"):
                fig = ax_or_fig.figure
            path = out_dir / filename
            fig.savefig(path, dpi=140, bbox_inches="tight", facecolor="white")
            plt.close("all")
            saved.append(path)
            print(f"Wrote {path}")
        except Exception as exc:
            print(f"Skip {filename}: {exc}")
            plt.close("all")

    return saved


def write_excel_report(
    path: Path,
    stats: pd.DataFrame,
    returns: pd.Series,
    benchmark: pd.Series,
    image_paths: list[Path],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PerfStats"
    ws["A1"] = "Pyfolio performance stats"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Relative-PE strategy vs {BENCH_LABEL}"
    ws["A3"] = f"Sample: {returns.index.min().date()} → {returns.index.max().date()} ({len(returns)} days)"

    # stats table starting row 5
    out = stats.copy()
    out.index.name = "metric"
    out = out.reset_index()
    for r_idx, row in enumerate(dataframe_to_rows(out, index=False, header=True), 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if not (isinstance(val, float) and (np.isnan(val))) else None)
            if r_idx == 5:
                cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22

    # Daily returns sheet
    ws_r = wb.create_sheet("DailyReturns")
    ret_df = pd.DataFrame({"strategy": returns, BENCH_LABEL: benchmark})
    ret_df.index.name = "date"
    ret_df = ret_df.reset_index()
    for r_idx, row in enumerate(dataframe_to_rows(ret_df, index=False, header=True), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws_r.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
            elif c_idx == 1:
                cell.number_format = "YYYY-MM-DD"
            else:
                cell.number_format = "0.00%"

    # Charts sheet with embedded images
    ws_c = wb.create_sheet("Charts")
    ws_c["A1"] = "Pyfolio charts"
    ws_c["A1"].font = Font(bold=True, size=14)
    row = 3
    for img_path in image_paths:
        if not img_path.exists():
            continue
        ws_c.cell(row=row, column=1, value=img_path.name).font = Font(bold=True)
        try:
            img = XLImage(str(img_path))
            # scale large tear sheets
            max_w = 1000
            if img.width and img.width > max_w:
                scale = max_w / img.width
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            ws_c.add_image(img, f"A{row + 1}")
            row += max(22, int((img.height or 400) / 18) + 3)
        except Exception as exc:
            ws_c.cell(row=row + 1, column=1, value=f"Could not embed {img_path.name}: {exc}")
            row += 3

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Wrote {path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Pyfolio tear sheet for Consumer Discretionary relative-PE")
    parser.add_argument("--sector", default=DEFAULT_SECTOR)
    parser.add_argument(
        "--xlsx",
        default=None,
        help="Path to portfolio_backtest.xlsx (default: data/sectors/consumer_discretionary/)",
    )
    args = parser.parse_args()

    out_dir = sector_dir(args.sector)
    xlsx = Path(args.xlsx) if args.xlsx else out_dir / "portfolio_backtest.xlsx"
    if not xlsx.exists():
        raise SystemExit(f"Missing {xlsx}. Run scripts/portfolio_backtest.py first.")

    bench_label = BENCH_LABEL
    pf_dir = out_dir / "pyfolio"
    pf_dir.mkdir(parents=True, exist_ok=True)

    print(f"Loading returns from {xlsx} ...")
    returns, benchmark = load_returns_from_excel(xlsx, bench_label=bench_label)
    print(f"Returns: {len(returns)} days ({returns.index.min().date()} → {returns.index.max().date()})")

    stats = compute_stats(returns, benchmark)
    stats_csv = pf_dir / "perf_stats.csv"
    stats.to_csv(stats_csv)
    print(f"Wrote {stats_csv}")
    print("\nKey stats:")
    for metric in [
        "Annual return",
        "Cumulative returns",
        "Annual volatility",
        "Sharpe ratio",
        "Calmar ratio",
        "Stability",
        "Max drawdown",
        "Sortino ratio",
        "Skew",
        "Kurtosis",
    ]:
        if metric in stats.index:
            s = stats.loc[metric]
            print(f"  {metric:22s}  strategy={s['strategy']!s:>12}  bench={s[bench_label]!s:>12}")

    images = save_tear_sheet_figures(returns, benchmark, pf_dir)
    report = out_dir / "pyfolio_analysis.xlsx"
    write_excel_report(report, stats, returns, benchmark, images)
    print("\nDone. Open:")
    print(f"  {report}")
    print(f"  {pf_dir}/")


if __name__ == "__main__":
    main()
