#!/usr/bin/env python3
"""$100k relative-PE/PB portfolio backtest vs S&P 500 sector index; Excel deliverable."""

from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src import DEFAULT_SECTOR, sector_dir
from src.pair_select import THRESHOLDS, select_top_significant_pairs, thresholds_for


INITIAL_CAPITAL = 100_000.0
TOP_N_PAIRS = 10
WARMUP = 126

SECTOR_BENCHMARKS: dict[str, dict[str, str]] = {
    "Consumer Discretionary": {
        "label": "S&P 500 Consumer Discretionary",
        "primary": "^SP500-25",
        "fallback": "XLY",
    },
}


def load_eligible_pairs(
    out_dir: Path,
    panel: pd.DataFrame,
    metric: str,
    top_n: int = TOP_N_PAIRS,
    score_min_days: int = 252,
    members: pd.DataFrame | None = None,
    significant: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Top-N same-subindustry significant pairs ranked by mean-reversion score (≤N → take all)."""
    if significant is None:
        sig_path = out_dir / f"significant_{metric}_pairs.csv"
        if not sig_path.exists():
            return pd.DataFrame()
        try:
            significant = pd.read_csv(sig_path)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()
    if significant is None or significant.empty:
        return pd.DataFrame()
    if members is None:
        members_path = out_dir / "members.csv"
        members = pd.read_csv(members_path) if members_path.exists() else None
    return select_top_significant_pairs(
        panel,
        significant,
        metric=metric,
        top_n=top_n,
        score_min_days=score_min_days,
        members=members,
    )


def _precompute_pair_series(
    panel: pd.DataFrame,
    close: pd.DataFrame,
    a: str,
    b: str,
    metric: str,
) -> pd.DataFrame | None:
    """Expanding z and leg returns for one pair; index = overlapping valid days."""
    suf = f"_{metric}"
    if a not in panel.columns or b not in panel.columns:
        return None
    if a not in close.columns or b not in close.columns:
        return None
    common = panel[[a, b]].join(close[[a, b]], lsuffix=suf, rsuffix="_px", how="inner")
    common = common.dropna()
    common = common[(common[f"{a}{suf}"] > 0) & (common[f"{b}{suf}"] > 0)]
    if len(common) < WARMUP + 2:
        return None
    spread = np.log(common[f"{a}{suf}"] / common[f"{b}{suf}"])
    mu = spread.expanding(min_periods=WARMUP).mean()
    sd = spread.expanding(min_periods=WARMUP).std(ddof=1)
    z = (spread - mu) / sd
    out = pd.DataFrame(
        {
            "z": z,
            "ret_a": common[f"{a}_px"].pct_change(),
            "ret_b": common[f"{b}_px"].pct_change(),
        },
        index=common.index,
    )
    return out


def simulate_sector_portfolio(
    panel: pd.DataFrame,
    close: pd.DataFrame,
    elig: pd.DataFrame,
    start: pd.Timestamp,
    capital: float,
    metric: str = "pe",
    max_slots: int = TOP_N_PAIRS,
) -> tuple[list[dict], pd.Series]:
    """
    Joint sector backtest: $capital split equally among currently open trades.

    Daily: accrue PnL at capital/n → exits → entries (score order, n < max_slots)
    → next day's n updates the equal split. Cash when n=0.
    """
    t = thresholds_for(metric)
    z_entry = t["z_entry"]
    z_exit = t["z_exit"]

    if elig is None or elig.empty:
        return [], pd.Series(dtype=float)

    pair_meta: list[dict] = []
    for rank, row in elig.reset_index(drop=True).iterrows():
        a, b = str(row["symbol_a"]), str(row["symbol_b"])
        series = _precompute_pair_series(panel, close, a, b, metric)
        if series is None:
            continue
        pair_meta.append(
            {
                "rank": int(rank),
                "symbol_a": a,
                "symbol_b": b,
                "pair": f"{a}-{b}",
                "score": float(row.get("score", np.nan)),
                "series": series,
            }
        )
    if not pair_meta:
        return [], pd.Series(dtype=float)

    # Score rank already from elig; keep that order for entry priority
    calendar = close.index.sort_values()
    calendar = calendar[calendar >= panel.index.min()]
    daily_pnl = pd.Series(0.0, index=calendar, dtype=float)

    # pid -> open state
    open_pos: dict[str, dict] = {}
    closed_trades: list[dict] = []

    def _finalize(st: dict, close_dt: pd.Timestamp, exit_z: float) -> None:
        notionals = st["notionals"]
        avg_notional = float(np.mean(notionals)) if notionals else 0.0
        pnl = float(st["pnl"])
        closed_trades.append(
            {
                "symbol_a": st["symbol_a"],
                "symbol_b": st["symbol_b"],
                "pair": st["pair"],
                "score": st["score"],
                "open_date": st["open_date"],
                "close_date": close_dt,
                "direction": st["direction"],
                "entry_z": st["entry_z"],
                "exit_z": exit_z,
                "hold_days": int(st["hold_days"]),
                "slot_notional": avg_notional,
                "pnl_usd": pnl,
                "return": (pnl / avg_notional) if avg_notional else np.nan,
            }
        )

    for i in range(1, len(calendar)):
        dt = calendar[i]
        n_open = len(open_pos)
        notional = (capital / n_open) if n_open else 0.0

        # 1) Accrue PnL for positions held overnight
        day_total = 0.0
        if n_open and dt >= start:
            for st in open_pos.values():
                ser = st["series"]
                if dt not in ser.index:
                    continue
                ra = ser.at[dt, "ret_a"]
                rb = ser.at[dt, "ret_b"]
                ra = float(ra) if np.isfinite(ra) else 0.0
                rb = float(rb) if np.isfinite(rb) else 0.0
                day_ret = 0.5 * st["pos_a"] * ra + 0.5 * st["pos_b"] * rb
                dollar = notional * day_ret
                st["pnl"] += dollar
                st["notionals"].append(notional)
                st["hold_days"] += 1
                day_total += dollar
        daily_pnl.iloc[i] = day_total

        # 2) Exits
        to_close: list[str] = []
        for pid, st in open_pos.items():
            ser = st["series"]
            if dt not in ser.index:
                continue
            zi = ser.at[dt, "z"]
            if np.isfinite(zi) and abs(float(zi)) <= z_exit:
                to_close.append(pid)
        for pid in to_close:
            st = open_pos.pop(pid)
            zi = st["series"].at[dt, "z"] if dt in st["series"].index else np.nan
            _finalize(st, dt, float(zi) if np.isfinite(zi) else np.nan)

        # 3) Entries (score order), only on/after trade start
        if dt >= start:
            for meta in pair_meta:
                if len(open_pos) >= max_slots:
                    break
                pid = meta["pair"]
                if pid in open_pos:
                    continue
                ser = meta["series"]
                if dt not in ser.index:
                    continue
                zi = ser.at[dt, "z"]
                if not np.isfinite(zi) or abs(float(zi)) < z_entry:
                    continue
                zi_f = float(zi)
                a, b = meta["symbol_a"], meta["symbol_b"]
                if zi_f > 0:
                    pos_a, pos_b = -1.0, 1.0
                    direction = f"short {a} / long {b}"
                else:
                    pos_a, pos_b = 1.0, -1.0
                    direction = f"long {a} / short {b}"
                open_pos[pid] = {
                    "symbol_a": a,
                    "symbol_b": b,
                    "pair": pid,
                    "score": meta["score"],
                    "pos_a": pos_a,
                    "pos_b": pos_b,
                    "direction": direction,
                    "entry_z": zi_f,
                    "open_date": dt,
                    "series": ser,
                    "pnl": 0.0,
                    "notionals": [],
                    "hold_days": 0,
                }

    # Force-close anything still open at the end
    if open_pos:
        last_dt = calendar[-1]
        for pid, st in list(open_pos.items()):
            zi = np.nan
            if last_dt in st["series"].index:
                zval = st["series"].at[last_dt, "z"]
                zi = float(zval) if np.isfinite(zval) else np.nan
            _finalize(st, last_dt, zi)
            del open_pos[pid]

    daily_pnl = daily_pnl.loc[daily_pnl.index >= start]
    return closed_trades, daily_pnl


def _download_close(ticker: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.Series:
    raw = yf.download(
        ticker,
        start=(start - pd.Timedelta(days=5)).strftime("%Y-%m-%d"),
        end=(end + pd.Timedelta(days=5)).strftime("%Y-%m-%d"),
        auto_adjust=True,
        progress=False,
    )
    if raw is None or raw.empty:
        raise RuntimeError(f"empty download for {ticker}")
    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"]
        if ticker in close.columns:
            px = close[ticker]
        else:
            px = close.iloc[:, 0]
    else:
        px = raw["Close"]
    px = px.copy()
    px.index = pd.to_datetime(px.index).tz_localize(None).normalize()
    px = px.sort_index().dropna()
    if px.empty:
        raise RuntimeError(f"no close prices for {ticker}")
    return px


def fetch_benchmark_equity(
    start: pd.Timestamp,
    end: pd.Timestamp,
    calendar: pd.DatetimeIndex,
    primary: str,
    fallback: str,
    capital: float = INITIAL_CAPITAL,
) -> tuple[pd.Series, str]:
    """$100k buy-and-hold in the sector index (primary), with ETF fallback."""
    used = primary
    try:
        px = _download_close(primary, start, end)
        print(f"Benchmark: using {primary}")
    except Exception as exc:
        print(f"Benchmark: {primary} failed ({exc}); falling back to {fallback}")
        px = _download_close(fallback, start, end)
        used = fallback

    aligned = px.reindex(calendar).ffill().bfill()
    base = float(aligned.iloc[0])
    equity = capital * (aligned / base)
    equity.name = "benchmark_value"
    return equity, used


def max_drawdown(equity: pd.Series) -> float:
    peak = equity.cummax()
    dd = equity / peak - 1.0
    return float(dd.min()) if len(dd) else np.nan


def render_performance_history_png(
    equity_df: pd.DataFrame, bench_label: str, strategy_label: str = "Relative-PE Strategy"
) -> bytes:
    """Performance History–style cumulative % chart with labeled axes."""
    dates = pd.to_datetime(equity_df["date"])
    strat_pct = equity_df["strategy_return_pct"].to_numpy(dtype=float) * 100.0
    bench_pct = equity_df["benchmark_return_pct"].to_numpy(dtype=float) * 100.0

    fig, ax = plt.subplots(figsize=(11, 5.5), dpi=140)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    ax.plot(dates, strat_pct, color="#2F6FED", linewidth=1.8, label=strategy_label)
    ax.plot(dates, bench_pct, color="#2C3E50", linewidth=1.8, label=bench_label)

    ax.axhline(0.0, color="#9AA0A6", linewidth=0.9)
    ax.set_title("PERFORMANCE HISTORY", fontsize=14, fontweight="bold", color="#1A73E8", loc="left", pad=12)
    ax.set_xlabel("Date", fontsize=11)
    ax.set_ylabel("Cumulative return (%)", fontsize=11)

    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda y, _: f"{y:.1f}%"))
    ax.xaxis.set_major_locator(mdates.YearLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    ax.xaxis.set_minor_locator(mdates.MonthLocator((1, 4, 7, 10)))
    fig.autofmt_xdate(rotation=0, ha="center")

    ax.grid(True, which="major", color="#E6E8EB", linewidth=0.8)
    ax.grid(True, which="minor", color="#F3F4F6", linewidth=0.5)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.spines["left"].set_color("#CFD4DA")
    ax.spines["bottom"].set_color("#CFD4DA")

    ax.legend(loc="lower right", frameon=True, fancybox=False, edgecolor="#E6E8EB")
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def write_excel(
    path: Path,
    summary_rows: list[tuple[str, object]],
    pairs_df: pd.DataFrame,
    equity_df: pd.DataFrame,
    trades_df: pd.DataFrame,
    chart_png: bytes,
    bench_label: str,
    strategy_label: str = "Relative-PE Strategy",
) -> None:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"{strategy_label} Portfolio Backtest"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Strategy vs {bench_label} buy-and-hold"
    row = 4
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        if isinstance(value, float) and ("return" in label.lower() or "drawdown" in label.lower()):
            cell.number_format = "0.00%"
        elif isinstance(value, float) and (
            "value" in label.lower() or "capital" in label.lower() or "pnl" in label.lower() or "notional" in label.lower()
        ):
            cell.number_format = "#,##0.00"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=f"Traded pairs (top {TOP_N_PAIRS} significant by score)").font = Font(bold=True)
    row += 1
    headers = list(pairs_df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h).font = Font(bold=True)
    for i, rec in enumerate(pairs_df.itertuples(index=False), 1):
        for c, val in enumerate(rec, 1):
            ws.cell(row=row + i, column=c, value=val)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22

    # Embed Performance History PNG
    img_path = path.parent / "_performance_history_tmp.png"
    img_path.write_bytes(chart_png)
    img = XLImage(str(img_path))
    img.width = 880
    img.height = 440
    ws.add_image(img, "D4")

    # --- Equity ---
    ws_eq = wb.create_sheet("Equity")
    eq = equity_df.copy()
    eq["date"] = pd.to_datetime(eq["date"]).dt.tz_localize(None)
    # Friendly headers for Excel chart
    rename = {
        "strategy_return_pct": f"{strategy_label} (%)",
        "benchmark_return_pct": f"{bench_label} (%)",
    }
    eq_out = eq.rename(columns={k: v for k, v in rename.items() if k in eq.columns})
    display = eq_out.copy()
    for r_idx, row_data in enumerate(dataframe_to_rows(display, index=False, header=True), 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_eq.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
                continue
            col_name = display.columns[c_idx - 1]
            if col_name == "date":
                cell.number_format = "YYYY-MM-DD"
            elif col_name in ("strategy_value", "benchmark_value"):
                cell.number_format = "#,##0.00"
            elif "(%)" in str(col_name) or col_name.endswith("_pct"):
                cell.number_format = "0.00%"

    n_eq = len(display) + 1
    # Column order: date, strategy_value, benchmark_value, strategy (%), sector index (%)
    chart = LineChart()
    chart.title = "PERFORMANCE HISTORY"
    chart.style = 10
    chart.y_axis.title = "Cumulative return (%)"
    chart.x_axis.title = "Date"
    chart.height = 12
    chart.width = 22
    data = Reference(ws_eq, min_col=4, min_row=1, max_col=5, max_row=n_eq)
    cats = Reference(ws_eq, min_col=1, min_row=2, max_row=n_eq)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_eq.add_chart(chart, "G2")

    for col, width in zip("ABCDE", [14, 16, 18, 24, 26]):
        ws_eq.column_dimensions[col].width = width

    # --- Trades ---
    ws_tr = wb.create_sheet("Trades")
    tr = trades_df.copy()
    if not tr.empty:
        for col in ("open_date", "close_date"):
            if col in tr.columns:
                tr[col] = pd.to_datetime(tr[col]).dt.tz_localize(None)
    for r_idx, row_data in enumerate(dataframe_to_rows(tr, index=False, header=True), 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws_tr.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
                continue
            col_name = tr.columns[c_idx - 1] if c_idx - 1 < len(tr.columns) else ""
            if col_name in ("open_date", "close_date"):
                cell.number_format = "YYYY-MM-DD"
            elif col_name in ("pnl_usd", "slot_notional"):
                cell.number_format = "#,##0.00"
            elif col_name in ("return",):
                cell.number_format = "0.00%"
            elif col_name in ("entry_z", "exit_z"):
                cell.number_format = "0.000"

    for col in ws_tr.columns:
        ws_tr.column_dimensions[col[0].column_letter].width = 16

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    try:
        img_path.unlink(missing_ok=True)
    except Exception:
        pass


def main() -> None:
    parser = argparse.ArgumentParser(description="Consumer Discretionary relative-PE portfolio backtest")
    parser.add_argument("--sector", default=DEFAULT_SECTOR)
    parser.add_argument("--capital", type=float, default=INITIAL_CAPITAL)
    parser.add_argument("--top-n", type=int, default=TOP_N_PAIRS, help="Max significant pairs to trade")
    args = parser.parse_args()
    metric = "pe"
    strategy_label = "Relative-PE Strategy"
    t = thresholds_for(metric)

    if args.sector not in SECTOR_BENCHMARKS:
        known = ", ".join(sorted(SECTOR_BENCHMARKS))
        raise SystemExit(f"Unsupported sector {args.sector!r}. Use: {known}")

    bench_cfg = SECTOR_BENCHMARKS[args.sector]
    bench_label = bench_cfg["label"]

    out_dir = sector_dir(args.sector)
    out_dir.mkdir(parents=True, exist_ok=True)
    panel = pd.read_csv(out_dir / "pe.csv", index_col=0, parse_dates=True)
    close = pd.read_csv(out_dir / "close.csv", index_col=0, parse_dates=True)
    members = pd.read_csv(out_dir / "members.csv")
    panel.index = pd.to_datetime(panel.index).tz_localize(None).normalize()
    close.index = pd.to_datetime(close.index).tz_localize(None).normalize()

    elig = load_eligible_pairs(
        out_dir,
        panel,
        metric=metric,
        top_n=args.top_n,
        score_min_days=252,
        members=members,
    )

    traded_path = out_dir / f"top{args.top_n}_significant_pe_pairs.csv"
    elig.to_csv(traded_path, index=False)
    print(f"Traded pairs -> {traded_path} ({len(elig)} rows)")

    start = max(pd.Timestamp("2021-01-01"), close.index.min())
    end = close.index.max()
    calendar = close.loc[close.index >= start].index

    n_pairs = len(elig)

    print(f"Sector: {args.sector} | metric=pe")
    print(
        f"Eligible pairs: {n_pairs} (max concurrent slots={args.top_n}); "
        f"allocation=equal among active trades; capital=${args.capital:,.0f}; "
        f"z_entry={t['z_entry']}; z_exit={t['z_exit']}; "
        f"trade_start={start.date()}; end={end.date()}"
    )

    if n_pairs == 0:
        print("  No scorable significant pairs; writing flat (cash) strategy returns.")
        all_trades: list[dict] = []
        daily_total = pd.Series(0.0, index=calendar, dtype=float)
    else:
        for _, row in elig.iterrows():
            print(f"  Eligible {row['symbol_a']}-{row['symbol_b']} (score={row['score']:.3f})")
        all_trades, daily_total = simulate_sector_portfolio(
            panel,
            close,
            elig,
            start=start,
            capital=args.capital,
            metric=metric,
            max_slots=args.top_n,
        )
        daily_total = daily_total.reindex(calendar).fillna(0.0)

    strategy = args.capital + daily_total.cumsum()
    strategy.name = "strategy_value"

    # Daily simple returns for multi-sector combiner (independent of absolute capital)
    daily_ret = strategy.pct_change().fillna(0.0)
    daily_ret.name = "strategy_return"
    ret_out = pd.DataFrame({"date": calendar, "strategy_return": daily_ret.reindex(calendar).fillna(0.0).to_numpy()})
    ret_path = out_dir / "strategy_daily_returns.csv"
    ret_out.to_csv(ret_path, index=False)
    print(f"Wrote {ret_path}")

    benchmark, bench_ticker = fetch_benchmark_equity(
        start,
        end,
        calendar,
        primary=bench_cfg["primary"],
        fallback=bench_cfg["fallback"],
        capital=args.capital,
    )

    strat_vals = strategy.reindex(calendar).to_numpy(dtype=float)
    bench_vals = benchmark.reindex(calendar).to_numpy(dtype=float)
    equity_df = pd.DataFrame(
        {
            "date": calendar,
            "strategy_value": strat_vals,
            "benchmark_value": bench_vals,
            "strategy_return_pct": strat_vals / args.capital - 1.0,
            "benchmark_return_pct": bench_vals / args.capital - 1.0,
        }
    )

    trades_df = pd.DataFrame(all_trades)
    if not trades_df.empty:
        trades_df = trades_df.sort_values(["open_date", "pair"]).reset_index(drop=True)
        cols = [
            "pair",
            "symbol_a",
            "symbol_b",
            "score",
            "open_date",
            "close_date",
            "direction",
            "entry_z",
            "exit_z",
            "hold_days",
            "slot_notional",
            "pnl_usd",
            "return",
        ]
        trades_df = trades_df[cols]

    strat_final = float(equity_df["strategy_value"].iloc[-1])
    bench_final = float(equity_df["benchmark_value"].iloc[-1])
    strat_ret = strat_final / args.capital - 1.0
    bench_ret = bench_final / args.capital - 1.0
    mdd = max_drawdown(equity_df.set_index("date")["strategy_value"])

    if n_pairs:
        pairs_out = elig[
            ["symbol_a", "symbol_b", "score", "pearson_r", "adf_t", "half_life", "exc_ge_1_5", "revert_rate"]
        ].copy()
        if "security_a" in elig.columns:
            pairs_out.insert(2, "security_a", elig["security_a"])
            pairs_out.insert(3, "security_b", elig["security_b"])
    else:
        pairs_out = pd.DataFrame(
            columns=["symbol_a", "symbol_b", "score", "pearson_r", "adf_t", "half_life", "exc_ge_1_5", "revert_rate"]
        )

    summary_rows = [
        ("Sector", args.sector),
        ("Metric", "PE"),
        ("Initial capital (USD)", float(args.capital)),
        ("Pair selection", f"Top {args.top_n} same-subindustry significant pairs by score"),
        ("Eligible pairs", n_pairs),
        ("Max concurrent slots", int(args.top_n)),
        (
            "Capital allocation",
            "Equal dynamic split of sector capital among active trades (rebalance on entry/exit)",
        ),
        ("Trade start", start.strftime("%Y-%m-%d")),
        ("End date", end.strftime("%Y-%m-%d")),
        ("Pair universe", "same_gics_subindustry"),
        ("Entry rule", f"|z| >= {t['z_entry']}"),
        ("Exit rule", f"|z| <= {t['z_exit']}"),
        ("Z warm-up (days)", WARMUP),
        ("Benchmark label", bench_label),
        ("Benchmark ticker used", bench_ticker),
        ("Final strategy value (USD)", strat_final),
        (f"Final {bench_label} value (USD)", bench_final),
        ("Strategy total return", strat_ret),
        (f"{bench_label} total return", bench_ret),
        ("Strategy max drawdown", mdd),
        ("Number of trades", int(len(trades_df))),
        ("Total trade PnL (USD)", float(trades_df["pnl_usd"].sum()) if not trades_df.empty else 0.0),
    ]

    chart_png = render_performance_history_png(equity_df, bench_label, strategy_label=strategy_label)
    (out_dir / "performance_history.png").write_bytes(chart_png)

    out_path = out_dir / "portfolio_backtest.xlsx"
    # Excel Equity sheet: prefer clean columns (drop aliases)
    equity_xlsx = equity_df[
        ["date", "strategy_value", "benchmark_value", "strategy_return_pct", "benchmark_return_pct"]
    ].copy()
    write_excel(
        out_path,
        summary_rows,
        pairs_out,
        equity_xlsx,
        trades_df,
        chart_png,
        bench_label,
        strategy_label=strategy_label,
    )
    print(f"\nWrote {out_path}")
    print(f"Also wrote {out_dir / 'performance_history.png'}")
    print(
        f"Strategy: ${strat_final:,.2f} ({strat_ret:.1%}) | "
        f"{bench_label} ({bench_ticker}): ${bench_final:,.2f} ({bench_ret:.1%})"
    )
    print(f"Trades: {len(trades_df)} | Max DD: {mdd:.1%}")


if __name__ == "__main__":
    main()
